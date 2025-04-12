import openpyxl as xl
## Load the workbook and select the sheet
wb=xl.load_workbook('transactions.xlsx')
sheet=wb['Sheet1']


# # cell=sheet.cell(1,1)
# # print(cell.value)
# # print(sheet.max_row)  #no.of rows

# for row in range(2,sheet.max_row + 1):
#  cell=sheet.cell(row,3)
#  print(cell.value)

#  #add new column
#  corrected_price=cell.value*0.9
#  corrected_price_cell=sheet.cell(row,4)
#  corrected_price_cell.value=corrected_price

# wb.save('transactions.xlsx')

#Chart
from openpyxl.chart import BarChart,Reference

#Reference class to select range of valus
value = Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=4)


chart=BarChart()
chart.add_data(value)
sheet.add_chart(chart,'E1')
wb.save('transactions_with_chart.xlsx')







#CLEAR CODE

import openpyxl as xl
from openpyxl.chart import BarChart,Reference
def process_workbook(filename):
    ## Load the workbook and select the sheet
    wb=xl.load_workbook('filename')
    sheet=wb['Sheet1']


    # cell=sheet.cell(1,1)
    # print(cell.value)
    # print(sheet.max_row)  #no.of rows

    for row in range(2,sheet.max_row + 1):
     cell=sheet.cell(row,3)
     print(cell.value)

     #add new column
     corrected_price=cell.value*0.9
     corrected_price_cell=sheet.cell(row,4)
     corrected_price_cell.value=corrected_price

    wb.save('filename')



    #Reference class to select range of valus
    value = Reference(sheet,
            min_row=2,
            max_row=sheet.max_row,
            min_col=4,
            max_col=4)


    chart=BarChart()
    chart.add_data(value)
    sheet.add_chart(chart,'E1')
    wb.save('filename')


